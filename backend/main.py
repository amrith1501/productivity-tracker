"""Productivity Tracker backend.

Tasks are persisted to SQLite (`app.db`), scoped per supervisor. Each
supervisor sees only their own team's tasks; the table view defaults to
the last 10 days while the full history is retained for trend charts.

Files dropped into `backend/tasks_inbox/` are still picked up by a
background watcher and round-robin-assigned across all active workers;
each task is automatically attributed to that worker's supervisor.
"""
from __future__ import annotations

import asyncio
import csv
import io
import json
import os
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from pathlib import Path
from typing import Literal, Optional

from fastapi import Depends, FastAPI, HTTPException, Request, Response
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

import hashlib
import secrets as _secrets
import time as _time

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from auth import (
    clear_session_cookie,
    current_user,
    hash_password,
    issue_token,
    login_throttle_check,
    login_throttle_record_failure,
    login_throttle_reset,
    require_supervisor,
    require_worker,
    set_session_cookie,
    verify_password,
)
from db import (
    consume_reset_token,
    count_unread_notifications,
    create_notification,
    create_task,
    create_user,
    deactivate_worker,
    delete_task as db_delete_task,
    delete_tasks_by_ids,
    delete_tasks_for_supervisor,
    employees_for_manager,
    existing_external_ids,
    get_inbox_processed_filenames,
    get_processed_imports,
    get_rr_index,
    get_task,
    get_user_by_username,
    get_worker_id_by_employee,
    init_db,
    list_active_workers_with_manager,
    list_notifications,
    list_tasks_for_supervisor,
    list_tasks_for_supervisor_range,
    list_tasks_for_worker,
    list_workers_for_manager,
    mark_import_processed,
    mark_inbox_processed,
    mark_login,
    mark_notifications_read,
    open_task_ids_for_assignee,
    reset_tasks_for_reassignment,
    set_rr_index,
    stats_for_supervisor,
    store_reset_token,
    task_count,
    update_password,
    update_task_fields,
)

# Header aliases accepted in CSV imports. Keys are normalised (lowercase,
# spaces/underscores/hyphens stripped) so users can write the column names
# however they like.
CSV_TITLE_KEYS = {"taskname", "title", "task", "name"}
CSV_DESC_KEYS = {"description", "desc", "details", "notes"}
CSV_ASSIGNEE_KEYS = {
    "assignedemployeeid", "employeeid", "assignee", "assignedto",
    "assigned", "employee", "worker", "username",
}
CSV_TASKID_KEYS = {"taskid", "id"}

BASE = Path(__file__).parent
INBOX = BASE / "tasks_inbox"
STATE_FILE = BASE / "state.json"
EMPLOYEES_FILE = BASE / "employees.json"
INBOX.mkdir(exist_ok=True)

# Default window applied to GET /api/tasks (used by the supervisor and
# worker dashboards). Older tasks remain in the DB for trend charts and
# are reachable by passing `?days=<N>` or `?all=true`.
DEFAULT_TASK_DAYS = 10

Status = Literal["pending", "in_progress", "submitted", "approved"]


class Task(BaseModel):
    id: str
    title: str
    description: str = ""
    assignee: str
    status: Status = "pending"
    source_file: str = ""
    created_at: str
    started_at: Optional[str] = None
    submitted_at: Optional[str] = None
    approved_at: Optional[str] = None


class TaskUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    assignee: Optional[str] = None
    status: Optional[Status] = None


def load_employees() -> list[str]:
    if EMPLOYEES_FILE.exists():
        return json.loads(EMPLOYEES_FILE.read_text())
    default = ["Alice", "Bob", "Carol", "Dave"]
    EMPLOYEES_FILE.write_text(json.dumps(default, indent=2))
    return default


def _migrate_state_json_to_sqlite() -> None:
    """One-shot import of legacy `state.json` tasks into the new SQLite
    table. Each task is attributed to the supervisor that manages its
    assignee (matched by display name). Tasks whose assignee has no
    current manager are dropped — they would have been orphaned anyway
    under the new per-supervisor model.

    After a successful migration the file is renamed to
    `state.json.migrated` so subsequent restarts skip this path.
    """
    if not STATE_FILE.exists():
        return
    if task_count() > 0:
        # Tasks already exist in SQLite — don't double-import. Move the
        # legacy file aside so we don't keep checking it.
        STATE_FILE.rename(STATE_FILE.with_suffix(".json.migrated"))
        return
    try:
        legacy = json.loads(STATE_FILE.read_text())
    except (OSError, json.JSONDecodeError) as e:
        print(f"[migrate] could not read state.json: {e}")
        return

    legacy_tasks = legacy.get("tasks", {}) or {}

    # Transfer the bookkeeping lists *first*. If we don't, the inbox watcher
    # will re-ingest every file in `tasks_inbox/` on the next tick because
    # the new `inbox_processed` table is empty — that's how the legacy
    # `state.json` accidentally double-counted tasks during the first
    # SQLite migration.
    for name in legacy.get("processed_files", []) or []:
        try:
            mark_inbox_processed(name)
        except Exception as e:
            print(f"[migrate] could not mark inbox file {name}: {e}")

    # Map display name -> manager_id by walking the user table.
    assignee_to_supervisor: dict[str, int] = {}
    for w in list_active_workers_with_manager():
        if w["employee"]:
            assignee_to_supervisor.setdefault(w["employee"].lower(),
                                              w["manager_id"])

    # `processed_imports` was a flat list of absolute file paths in the
    # legacy state.json. The new schema scopes them per supervisor; we
    # don't know which supervisor imported each path historically, so we
    # apply each remembered path to every active supervisor as a safe
    # default — preventing re-import for everyone.
    legacy_import_paths = legacy.get("processed_imports", []) or []
    if legacy_import_paths:
        supervisor_ids = {sid for sid in assignee_to_supervisor.values()}
        for sid in supervisor_ids:
            for p in legacy_import_paths:
                try:
                    mark_import_processed(sid, p)
                except Exception as e:
                    print(f"[migrate] could not mark import {p}: {e}")

    if not isinstance(legacy_tasks, dict) or not legacy_tasks:
        STATE_FILE.rename(STATE_FILE.with_suffix(".json.migrated"))
        return

    imported = 0
    skipped = 0
    for tid, t in legacy_tasks.items():
        assignee = (t.get("assignee") or "").strip()
        sid = assignee_to_supervisor.get(assignee.lower())
        if not sid:
            skipped += 1
            continue
        try:
            create_task(
                task_id=str(tid),
                supervisor_id=sid,
                title=t.get("title", "Untitled"),
                description=t.get("description", "") or "",
                assignee=assignee,
                status=t.get("status", "pending"),
                source_file=t.get("source_file", "") or "",
                created_at=t.get("created_at",
                                 datetime.utcnow().isoformat()),
            )
            updates = {k: t.get(k) for k in
                       ("started_at", "submitted_at", "approved_at")
                       if t.get(k)}
            if updates:
                update_task_fields(str(tid), updates)
            imported += 1
        except Exception as e:
            print(f"[migrate] failed on task {tid}: {e}")
            skipped += 1

    print(f"[migrate] state.json -> SQLite: imported={imported} "
          f"skipped={skipped} (no matching supervisor)")
    STATE_FILE.rename(STATE_FILE.with_suffix(".json.migrated"))


EMPLOYEES = load_employees()
init_db()
_migrate_state_json_to_sqlite()


class LoginRequest(BaseModel):
    username: str
    password: str


class RegisterRequest(BaseModel):
    username: str
    password: str
    employee: str  # display name


class CreateWorkerRequest(BaseModel):
    username: str
    password: str
    employee: str


class ImportDirectoryRequest(BaseModel):
    path: str


class ResetRequestPayload(BaseModel):
    username: str


class ResetConfirmPayload(BaseModel):
    token: str
    new_password: str


class MarkReadPayload(BaseModel):
    ids: Optional[list[int]] = None  # None => mark all unread as read


def _norm_header(name: str) -> str:
    return "".join(ch for ch in name.lower() if ch.isalnum())


def _parse_csv(text: str) -> list[dict]:
    """Parse a CSV with flexible headers.

    Recognised columns (case/space/underscore-insensitive):
      - Task ID                    -> external_id (optional)
      - Task Name / Title          -> title (required)
      - Description / Details      -> description (optional)
      - Assigned Employee_ID       -> assignee_hint (optional username
                                     or display name; blank => auto)
    """
    # Strip a UTF-8 BOM if present (Excel adds one when saving CSVs).
    if text.startswith("\ufeff"):
        text = text.lstrip("\ufeff")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    header_map: dict[str, str] = {}
    for raw in reader.fieldnames:
        key = _norm_header(raw or "")
        if key in CSV_TITLE_KEYS:
            header_map[raw] = "title"
        elif key in CSV_DESC_KEYS:
            header_map[raw] = "description"
        elif key in CSV_ASSIGNEE_KEYS:
            header_map[raw] = "assignee_hint"
        elif key in CSV_TASKID_KEYS:
            header_map[raw] = "external_id"
    if "title" not in header_map.values():
        raise ValueError("CSV is missing a Task Name / Title column")
    out: list[dict] = []
    for row in reader:
        entry: dict = {"title": "", "description": ""}
        for raw, canonical in header_map.items():
            val = (row.get(raw) or "").strip()
            if val:
                entry[canonical] = val
        if entry.get("title"):
            out.append(entry)
    return out


def parse_task_file(path: Path) -> list[dict]:
    """Extract task entries from a daily file. Supports .json, .txt, .csv.

    Returned dicts always carry `title`/`description`. CSV rows may
    additionally carry `assignee_hint` (a username or display name to
    target) and `external_id` (the file's own Task ID column, kept for
    traceability).
    """
    text = path.read_text(encoding="utf-8").strip()
    if not text:
        return []
    suffix = path.suffix.lower()
    if suffix == ".json":
        data = json.loads(text)
        items = data.get("tasks", data) if isinstance(data, dict) else data
        out = []
        for it in items:
            if isinstance(it, str):
                out.append({"title": it, "description": ""})
            else:
                entry = {
                    "title": it.get("title", "Untitled"),
                    "description": it.get("description", ""),
                }
                # Accept any of the common spellings for an external Task ID.
                ext = (it.get("task_id") or it.get("id")
                       or it.get("taskId") or it.get("external_id"))
                if ext is not None and str(ext).strip():
                    entry["external_id"] = str(ext).strip()
                out.append(entry)
        return out
    if suffix == ".csv":
        return _parse_csv(text)
    # plain text: one task per non-empty line
    return [{"title": line.strip(), "description": ""}
            for line in text.splitlines() if line.strip()]


_inbox_rr_index = 0  # Process-local counter for the global inbox watcher.


def ingest_inbox() -> int:
    """Scan `tasks_inbox/` for new files and create tasks.

    Each task is round-robin-assigned across **active workers that have
    a supervisor** (`users.manager_id IS NOT NULL`). The task's
    `supervisor_id` is set to that worker's manager so it shows up only
    on the right supervisor's dashboard. CSV rows that name a specific
    worker (by username or display name) bypass the round-robin and go
    straight to that worker.

    Workers without a supervisor (e.g. self-registered accounts that
    haven't been claimed yet) are skipped. Files become inert once
    processed: their filenames are recorded in `inbox_processed`.
    """
    global _inbox_rr_index
    workers = list_active_workers_with_manager()
    if not workers:
        return 0

    by_username = {(w["username"] or "").lower(): w for w in workers}
    by_employee = {(w["employee"] or "").lower(): w for w in workers}
    by_userid = {str(w["id"]): w for w in workers}

    # Cache of external Task IDs already stored per supervisor, so a Task ID
    # is ingested at most once per supervisor even if it shows up again.
    seen_ext: dict[int, set[str]] = {}

    def _ext_seen(supervisor_id: int, ext: str) -> bool:
        if supervisor_id not in seen_ext:
            seen_ext[supervisor_id] = existing_external_ids(supervisor_id)
        return ext in seen_ext[supervisor_id]

    processed = get_inbox_processed_filenames()
    files = sorted(p for p in INBOX.iterdir() if p.is_file())
    new_count = 0
    dupe_count = 0

    for path in files:
        if path.name in processed:
            continue
        try:
            entries = parse_task_file(path)
        except Exception as e:
            print(f"[ingest] skip {path.name}: {e}")
            continue
        for entry in entries:
            hint = (entry.get("assignee_hint") or "").strip().lower()
            worker = (by_employee.get(hint)
                      or by_username.get(hint)
                      or by_userid.get(hint))
            if not worker:
                worker = workers[_inbox_rr_index % len(workers)]
                _inbox_rr_index += 1
            supervisor_id = worker["manager_id"]
            ext = entry.get("external_id")
            if ext and _ext_seen(supervisor_id, ext):
                dupe_count += 1
                continue
            tid = str(uuid.uuid4())
            create_task(
                task_id=tid,
                supervisor_id=supervisor_id,
                external_id=ext,
                title=entry["title"],
                description=entry.get("description", "") or "",
                assignee=worker["employee"],
                status="pending",
                source_file=path.name,
                created_at=datetime.utcnow().isoformat(),
            )
            create_notification(
                worker["id"], "task_assigned",
                f"New task assigned: {entry['title']}", tid,
            )
            if ext:
                seen_ext[supervisor_id].add(ext)
            new_count += 1
        mark_inbox_processed(path.name)
    if new_count or dupe_count:
        print(f"[ingest] added {new_count} tasks from inbox"
              f"{f', skipped {dupe_count} duplicate Task ID(s)' if dupe_count else ''}")
    return new_count


async def inbox_watcher():
    while True:
        try:
            ingest_inbox()
        except Exception as e:
            print(f"[watcher] error: {e}")
        await asyncio.sleep(5)


@asynccontextmanager
async def lifespan(app: FastAPI):
    ingest_inbox()
    task = asyncio.create_task(inbox_watcher())
    yield
    task.cancel()


app = FastAPI(lifespan=lifespan)

# CORS: allow only configured origins, and only with credentials (cookies).
# In dev, the Vite proxy makes the API same-origin, so this is mostly a
# safety net. In prod, set PT_ALLOWED_ORIGINS="https://your.app".
_origins = [o.strip() for o in
            os.environ.get("PT_ALLOWED_ORIGINS", "http://localhost:5173").split(",")
            if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=_origins,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PATCH", "DELETE", "OPTIONS"],
    allow_headers=["Content-Type"],
)


def _client_ip(request: Request) -> str:
    # If you deploy behind a trusted reverse proxy, configure
    # uvicorn with --proxy-headers to populate request.client correctly.
    return request.client.host if request.client else "unknown"


@app.post("/api/login")
def login(req: LoginRequest, request: Request, response: Response):
    ip = _client_ip(request)
    login_throttle_check(ip)

    user = get_user_by_username(req.username)
    # Always run pbkdf2 — even on unknown user — to keep response time
    # roughly constant and avoid leaking which usernames exist.
    if user is None:
        verify_password(req.password,
                        expected_hash=b"\0" * 32,
                        salt=b"\0" * 16,
                        iterations=600_000)
        login_throttle_record_failure(ip)
        raise HTTPException(401, "Invalid credentials")

    ok = verify_password(req.password,
                         expected_hash=user["password_hash"],
                         salt=user["salt"],
                         iterations=user["iterations"])
    if not ok:
        login_throttle_record_failure(ip)
        raise HTTPException(401, "Invalid credentials")

    login_throttle_reset(ip)
    mark_login(user["id"])
    token = issue_token(
        user_id=user["id"],
        username=user["username"],
        role=user["role"],
        employee=user["employee"],
    )
    set_session_cookie(response, token)
    return {"user": {"username": user["username"], "role": user["role"],
                     "employee": user["employee"]}}


@app.post("/api/logout")
def logout(response: Response):
    clear_session_cookie(response)
    return {"ok": True}


@app.post("/api/register", status_code=201)
def register(req: RegisterRequest, response: Response):
    """Self-service registration. Always creates a 'worker' account.
    Supervisor accounts must be provisioned via the CLI for safety.
    """
    uname = req.username.strip().lower()
    employee = req.employee.strip()
    if len(uname) < 3 or not uname.replace("_", "").isalnum():
        raise HTTPException(400, "Username must be 3+ chars, alphanumeric/underscore.")
    if not employee:
        raise HTTPException(400, "Display name is required.")
    if get_user_by_username(uname):
        raise HTTPException(409, "Username already taken.")
    try:
        digest, salt, iters = hash_password(req.password)
    except ValueError as e:
        raise HTTPException(400, str(e))
    uid = create_user(uname, digest, salt, iters, "worker", employee)
    # Add new employee to roster if not already present.
    if employee not in EMPLOYEES:
        EMPLOYEES.append(employee)
        EMPLOYEES_FILE.write_text(json.dumps(EMPLOYEES, indent=2))
    token = issue_token(user_id=uid, username=uname, role="worker",
                        employee=employee)
    set_session_cookie(response, token)
    return {"user": {"username": uname, "role": "worker", "employee": employee}}


PASSWORD_RESET_TTL = 30 * 60  # 30 minutes


@app.post("/api/password-reset/request")
def request_password_reset(payload: ResetRequestPayload, request: Request):
    """Generate a single-use reset token.

    Always returns the same response to avoid leaking which usernames
    exist. The token is delivered out-of-band — in production hook this
    up to your email/SMS provider. Here it's logged to the server
    console for the admin to deliver.
    """
    login_throttle_check(_client_ip(request))  # reuse throttle to slow abuse
    user = get_user_by_username(payload.username.strip().lower())
    if user:
        raw = _secrets.token_urlsafe(32)
        token_hash = hashlib.sha256(raw.encode()).digest()
        store_reset_token(token_hash, user["id"],
                          int(_time.time()) + PASSWORD_RESET_TTL)
        print(f"[password-reset] user={user['username']} token={raw} "
              f"(expires in {PASSWORD_RESET_TTL // 60} min)")
    return {"ok": True,
            "message": "If that account exists, a reset token has been sent."}


@app.post("/api/password-reset/confirm")
def confirm_password_reset(payload: ResetConfirmPayload):
    token_hash = hashlib.sha256(payload.token.encode()).digest()
    user_id = consume_reset_token(token_hash)
    if not user_id:
        raise HTTPException(400, "Invalid or expired reset token.")
    try:
        digest, salt, iters = hash_password(payload.new_password)
    except ValueError as e:
        raise HTTPException(400, str(e))
    update_password(user_id, digest, salt, iters)
    return {"ok": True}


@app.get("/api/me")
def me(user: dict = Depends(current_user)):
    return {"username": user["sub"], "role": user["role"],
            "employee": user.get("employee")}


@app.get("/api/notifications")
def get_notifications(user: dict = Depends(current_user)):
    """Current user's recent alerts plus their unread count."""
    return {
        "items": list_notifications(user["uid"]),
        "unread": count_unread_notifications(user["uid"]),
    }


@app.post("/api/notifications/read")
def read_notifications(payload: MarkReadPayload,
                       user: dict = Depends(current_user)):
    """Mark notifications as read. Pass specific `ids`, or omit to clear
    all of the caller's unread alerts.
    """
    marked = mark_notifications_read(user["uid"], payload.ids)
    return {"marked": marked, "unread": count_unread_notifications(user["uid"])}


@app.get("/api/supervisor/workers")
def list_my_workers(user: dict = Depends(require_supervisor)):
    return list_workers_for_manager(user["uid"])


@app.delete("/api/supervisor/workers/{worker_id}")
def remove_worker(worker_id: int,
                  delete_tasks: bool = False,
                  user: dict = Depends(require_supervisor)):
    """Remove a worker from the supervisor's team.

    The worker is deactivated (soft-delete) so historical task records keep
    a meaningful assignee. By default their open tasks are returned to the
    pool as `pending` and reassigned round-robin across the remaining team;
    pass `delete_tasks=true` to drop them instead.
    """
    row = deactivate_worker(worker_id, user["uid"])
    if not row:
        # Either the worker doesn't exist, isn't on this team, or is already
        # inactive. We treat all three the same to avoid leaking membership.
        raise HTTPException(404, "Worker not found on your team.")

    employee_name = row.get("employee") or ""
    tasks_deleted = 0
    tasks_reassigned = 0
    remaining_team = _team_employees(user)

    if employee_name:
        affected_ids = open_task_ids_for_assignee(employee_name, user["uid"])
        if delete_tasks:
            tasks_deleted = delete_tasks_by_ids(affected_ids)
        elif remaining_team and affected_ids:
            rr = get_rr_index(user["uid"])
            # Group the IDs by their newly-chosen assignee so we can hand
            # them to the bulk-reassign helper in one shot per assignee.
            buckets: dict[str, list[str]] = {}
            for tid in affected_ids:
                new_assignee = remaining_team[rr % len(remaining_team)]
                rr += 1
                buckets.setdefault(new_assignee, []).append(tid)
            for new_assignee, ids in buckets.items():
                tasks_reassigned += reset_tasks_for_reassignment(ids, new_assignee)
                worker_uid = get_worker_id_by_employee(user["uid"], new_assignee)
                if worker_uid:
                    create_notification(
                        worker_uid, "task_assigned",
                        f"{len(ids)} task(s) reassigned to you from "
                        f"{employee_name}", None,
                    )
            set_rr_index(user["uid"], rr)
        # If no team left and not deleting, tasks stay assigned to the
        # removed worker's display name as a historical record.

    return {
        "ok": True,
        "removed": {"id": worker_id, "username": row.get("username"),
                    "employee": employee_name},
        "tasks_deleted": tasks_deleted,
        "tasks_reassigned": tasks_reassigned,
    }


@app.post("/api/supervisor/workers", status_code=201)
def add_worker(req: CreateWorkerRequest,
               user: dict = Depends(require_supervisor)):
    uname = req.username.strip().lower()
    employee = req.employee.strip()
    if len(uname) < 3 or not uname.replace("_", "").isalnum():
        raise HTTPException(400, "Username must be 3+ chars, alphanumeric/underscore.")
    if not employee:
        raise HTTPException(400, "Display name is required.")
    if get_user_by_username(uname):
        raise HTTPException(409, "Username already taken.")
    try:
        digest, salt, iters = hash_password(req.password)
    except ValueError as e:
        raise HTTPException(400, str(e))
    uid = create_user(uname, digest, salt, iters, "worker", employee,
                      manager_id=user["uid"])
    if employee not in EMPLOYEES:
        EMPLOYEES.append(employee)
        EMPLOYEES_FILE.write_text(json.dumps(EMPLOYEES, indent=2))
    return {"id": uid, "username": uname, "employee": employee}


def _team_employees(user: dict) -> list[str]:
    """Display names of the supervisor's direct reports."""
    return employees_for_manager(user["uid"])


@app.get("/api/employees")
def get_employees(user: dict = Depends(current_user)):
    # Supervisors get only their own team; workers get the full roster
    # for any UI that needs it (currently none, but keep the contract).
    if user["role"] == "supervisor":
        return _team_employees(user)
    return EMPLOYEES


@app.get("/api/tasks")
def list_tasks(status: Optional[Status] = None,
               assignee: Optional[str] = None,
               days: Optional[int] = None,
               all: bool = False,
               user: dict = Depends(current_user)):
    """List tasks visible to the caller.

    By default the response is limited to tasks `created_at` within the
    last `DEFAULT_TASK_DAYS` (10) days so the dashboard stays focused on
    recent work. Pass `?days=<N>` to widen the window or `?all=true` to
    return the full history (used by the productivity trend charts).
    """
    window = None if all else (days if days and days > 0 else DEFAULT_TASK_DAYS)
    if user["role"] == "worker":
        items = list_tasks_for_worker(
            user.get("employee") or "",
            days=window,
            status=status,
        )
    else:
        items = list_tasks_for_supervisor(
            user["uid"],
            days=window,
            status=status,
            assignee=assignee,
        )
    return items


@app.get("/api/tasks/{task_id}")
def get_one_task(task_id: str, user: dict = Depends(current_user)):
    t = get_task(task_id)
    if not t:
        raise HTTPException(404, "Task not found")
    if user["role"] == "worker":
        if t["assignee"] != user.get("employee"):
            raise HTTPException(403, "Not your task")
    else:
        if t["supervisor_id"] != user["uid"]:
            raise HTTPException(403, "Not your team's task")
    return t


@app.patch("/api/tasks/{task_id}")
def update_task(task_id: str, payload: TaskUpdate,
                user: dict = Depends(require_supervisor)):
    t = get_task(task_id)
    if not t:
        raise HTTPException(404, "Task not found")
    if t["supervisor_id"] != user["uid"]:
        raise HTTPException(403, "Not your team's task")
    data = payload.model_dump(exclude_unset=True)
    reassigned_to = None
    if "assignee" in data:
        team = set(_team_employees(user))
        if data["assignee"] not in team:
            raise HTTPException(400, "Assignee must be on your team")
        if data["assignee"] != t["assignee"]:
            reassigned_to = data["assignee"]
    updated = update_task_fields(task_id, data)
    if reassigned_to:
        worker_uid = get_worker_id_by_employee(user["uid"], reassigned_to)
        if worker_uid:
            create_notification(
                worker_uid, "task_assigned",
                f"New task assigned: {updated['title']}", task_id,
            )
    return updated


@app.post("/api/tasks/{task_id}/start")
def start_task(task_id: str, user: dict = Depends(require_worker)):
    t = get_task(task_id)
    if not t:
        raise HTTPException(404, "Task not found")
    if t["assignee"] != user.get("employee"):
        raise HTTPException(403, "Not your task")
    if t["status"] != "pending":
        raise HTTPException(400, f"Cannot start from status {t['status']}")
    return update_task_fields(task_id, {
        "status": "in_progress",
        "started_at": datetime.utcnow().isoformat(),
    })


@app.post("/api/tasks/{task_id}/submit")
def submit_task(task_id: str, user: dict = Depends(require_worker)):
    t = get_task(task_id)
    if not t:
        raise HTTPException(404, "Task not found")
    if t["assignee"] != user.get("employee"):
        raise HTTPException(403, "Not your task")
    if t["status"] != "in_progress":
        raise HTTPException(400, f"Cannot submit from status {t['status']}")
    updated = update_task_fields(task_id, {
        "status": "submitted",
        "submitted_at": datetime.utcnow().isoformat(),
    })
    # Alert the owning supervisor that work is ready for review.
    create_notification(
        t["supervisor_id"], "task_submitted",
        f"{t['assignee']} submitted: {t['title']}", task_id,
    )
    return updated


@app.post("/api/tasks/{task_id}/approve")
def approve_task(task_id: str, user: dict = Depends(require_supervisor)):
    t = get_task(task_id)
    if not t:
        raise HTTPException(404, "Task not found")
    if t["supervisor_id"] != user["uid"]:
        raise HTTPException(403, "Not your team's task")
    if t["status"] != "submitted":
        raise HTTPException(400, f"Cannot approve from status {t['status']}")
    return update_task_fields(task_id, {
        "status": "approved",
        "approved_at": datetime.utcnow().isoformat(),
    })


@app.post("/api/ingest")
def force_ingest(user: dict = Depends(require_supervisor)):
    return {"added": ingest_inbox()}


@app.delete("/api/tasks")
def delete_tasks(status: Optional[Status] = None,
                 user: dict = Depends(require_supervisor)):
    """Bulk-delete tasks belonging to the supervisor's team.

    Optional `status` filter limits deletion (e.g. `status=approved` to
    archive cleanup). Without it, every task owned by the calling
    supervisor is removed.
    """
    deleted = delete_tasks_for_supervisor(user["uid"], status=status)
    return {"deleted": deleted, "status": status}


@app.delete("/api/tasks/{task_id}")
def delete_single_task(task_id: str, user: dict = Depends(require_supervisor)):
    """Delete one task. Only the owning supervisor may remove it."""
    t = get_task(task_id)
    if not t:
        raise HTTPException(404, "Task not found")
    if t["supervisor_id"] != user["uid"]:
        raise HTTPException(403, "Not your team's task")
    db_delete_task(task_id)
    return {"ok": True, "id": task_id}


def _build_assignee_lookup(supervisor_uid: int,
                           team: list[str]) -> dict[str, str]:
    """Map every accepted spelling of an assignee to the canonical display
    name. Keys are lowercased; values are the `employee` column as stored
    in the DB. Supports matching by username, display name, or numeric
    user-id (so CSVs that say `Employee_ID=42` still resolve).
    """
    lookup: dict[str, str] = {}
    for w in list_workers_for_manager(supervisor_uid):
        if w.get("employee"):
            lookup[w["employee"].lower()] = w["employee"]
        if w.get("username"):
            lookup[w["username"].lower()] = w["employee"] or w["username"]
        if w.get("id") is not None:
            lookup[str(w["id"])] = w["employee"] or w["username"]
    # Also include any display names that came from `team` but weren't
    # surfaced as a worker row (defensive — keeps behaviour identical for
    # the legacy default roster).
    for name in team:
        lookup.setdefault(name.lower(), name)
    return lookup


def _ingest_directory_for_team(directory: Path, team: list[str],
                               supervisor_uid: int) -> dict:
    """Parse all .json/.txt/.csv files in `directory` and create tasks
    owned by `supervisor_uid`. CSV rows with an `Assigned Employee_ID`
    that matches a team member go straight to that worker; everything
    else is assigned round-robin within the supervisor's own team using
    a persisted per-supervisor counter.
    """
    added = 0
    skipped = 0
    duplicates = 0
    errors: list[str] = []
    processed = get_processed_imports(supervisor_uid)
    rr = get_rr_index(supervisor_uid)
    assignee_lookup = _build_assignee_lookup(supervisor_uid, team)
    # Existing Task IDs for this supervisor — used to skip rows that have
    # already been imported under the same Task ID, even from a new file.
    seen_ext = existing_external_ids(supervisor_uid)
    # Resolve display name -> worker user id once, to alert the assignee.
    worker_id_by_employee = {
        w["employee"]: w["id"]
        for w in list_workers_for_manager(supervisor_uid)
        if w.get("employee")
    }

    files = sorted(p for p in directory.iterdir()
                   if p.is_file()
                   and p.suffix.lower() in (".json", ".txt", ".csv"))
    for path in files:
        abs_path = str(path.resolve())
        if abs_path in processed:
            skipped += 1
            continue
        try:
            entries = parse_task_file(path)
        except Exception as e:
            errors.append(f"{path.name}: {e}")
            continue
        for entry in entries:
            ext = entry.get("external_id")
            if ext and ext in seen_ext:
                duplicates += 1
                continue
            hint = (entry.get("assignee_hint") or "").strip().lower()
            assignee = assignee_lookup.get(hint) if hint else None
            if hint and not assignee:
                errors.append(
                    f"{path.name}: unknown employee '{entry['assignee_hint']}' "
                    f"for task '{entry['title']}' — auto-assigned instead")
            if not assignee:
                assignee = team[rr % len(team)]
                rr += 1
            tid = str(uuid.uuid4())
            create_task(
                task_id=tid,
                supervisor_id=supervisor_uid,
                external_id=ext,
                title=entry["title"],
                description=entry.get("description", "") or "",
                assignee=assignee,
                status="pending",
                source_file=path.name,
                created_at=datetime.utcnow().isoformat(),
            )
            assignee_uid = worker_id_by_employee.get(assignee)
            if assignee_uid:
                create_notification(
                    assignee_uid, "task_assigned",
                    f"New task assigned: {entry['title']}", tid,
                )
            if ext:
                seen_ext.add(ext)
            added += 1
        mark_import_processed(supervisor_uid, abs_path)
        processed.add(abs_path)

    set_rr_index(supervisor_uid, rr)
    return {"added": added, "skipped": skipped, "duplicates": duplicates,
            "errors": errors, "scanned": len(files)}


@app.post("/api/supervisor/import-directory")
def import_directory(req: ImportDirectoryRequest,
                     user: dict = Depends(require_supervisor)):
    """One-shot import of all task files from a server-side directory.
    Tasks are assigned round-robin across the supervisor's own team.
    """
    raw = (req.path or "").strip()
    if not raw:
        raise HTTPException(400, "Path is required.")
    try:
        directory = Path(raw).expanduser().resolve()
    except (OSError, RuntimeError) as e:
        raise HTTPException(400, f"Invalid path: {e}")
    if not directory.exists():
        raise HTTPException(400, "Path does not exist on the server.")
    if not directory.is_dir():
        raise HTTPException(400, "Path is not a directory.")

    team = _team_employees(user)
    if not team:
        raise HTTPException(400, "You have no workers on your team yet. "
                                 "Add a worker before importing tasks.")

    result = _ingest_directory_for_team(directory, team, user["uid"])
    result["path"] = str(directory)
    return result


@app.get("/api/stats")
def stats(user: dict = Depends(require_supervisor)):
    agg = stats_for_supervisor(user["uid"])
    # Make sure every active team member shows up in `by_assignee`, even
    # workers that haven't received a task yet, so the workload card on
    # the dashboard renders consistently.
    for emp in _team_employees(user):
        agg["by_assignee"].setdefault(emp, {})
    return agg


def _parse_date_bound(value: Optional[str], *, end: bool) -> Optional[str]:
    """Validate a YYYY-MM-DD query param and turn it into an ISO bound that
    matches how `created_at` is stored. The end bound is pushed to the end
    of the day so the whole day is included.
    """
    if not value:
        return None
    try:
        d = datetime.strptime(value, "%Y-%m-%d")
    except ValueError:
        raise HTTPException(400, f"Invalid date '{value}'. Use YYYY-MM-DD.")
    if end:
        return d.strftime("%Y-%m-%dT23:59:59.999999")
    return d.strftime("%Y-%m-%dT00:00:00")


def _fmt_dt(value: Optional[str]) -> str:
    """Render an ISO timestamp as 'YYYY-MM-DD HH:MM' for the spreadsheet."""
    if not value:
        return ""
    try:
        return datetime.fromisoformat(value).strftime("%Y-%m-%d %H:%M")
    except ValueError:
        return value


_STATUS_ORDER = ["pending", "in_progress", "submitted", "approved"]


def _build_export_workbook(tasks: list[dict], team: list[str], *,
                           start: Optional[str], end: Optional[str]) -> Workbook:
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0A1F44")
    wrap = Alignment(vertical="top", wrap_text=True)

    def _style_header(ws, ncols):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="center")
        ws.freeze_panes = "A2"

    # ---- Sheet 1: Tasks ----
    wb = Workbook()
    ws = wb.active
    ws.title = "Tasks"
    task_headers = ["Task ID", "Title", "Description", "Assignee", "Status",
                    "Source File", "Created", "Started", "Submitted",
                    "Approved"]
    ws.append(task_headers)
    for t in tasks:
        ws.append([
            t.get("external_id") or "",
            t.get("title") or "",
            t.get("description") or "",
            t.get("assignee") or "",
            (t.get("status") or "").replace("_", " "),
            t.get("source_file") or "",
            _fmt_dt(t.get("created_at")),
            _fmt_dt(t.get("started_at")),
            _fmt_dt(t.get("submitted_at")),
            _fmt_dt(t.get("approved_at")),
        ])
    widths = [12, 34, 44, 16, 13, 18, 17, 17, 17, 17]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            cell.alignment = wrap
    _style_header(ws, len(task_headers))

    # ---- Sheet 2: Productivity ----
    ps = wb.create_sheet("Productivity")
    period = (f"{start[:10] if start else 'beginning'} "
              f"to {end[:10] if end else 'today'}")
    ps.append([f"Productivity summary ({period})"])
    ps["A1"].font = Font(bold=True, size=13)
    ps.append([])
    prod_headers = ["Employee", "Total", "Pending", "In progress",
                    "Submitted", "Approved", "Completion rate",
                    "Avg turnaround (hrs)"]
    ps.append(prod_headers)
    header_row = ps.max_row

    # Aggregate from the filtered task set so the summary matches the range.
    agg: dict[str, dict] = {e: {s: 0 for s in _STATUS_ORDER} for e in team}
    cycle: dict[str, list[float]] = {e: [] for e in team}
    for t in tasks:
        a = t.get("assignee") or ""
        bucket = agg.setdefault(a, {s: 0 for s in _STATUS_ORDER})
        bucket[t.get("status", "pending")] = bucket.get(t.get("status"), 0) + 1
        if t.get("status") == "approved" and t.get("started_at") and t.get("approved_at"):
            try:
                hrs = (datetime.fromisoformat(t["approved_at"])
                       - datetime.fromisoformat(t["started_at"])).total_seconds() / 3600
                if hrs >= 0:
                    cycle.setdefault(a, []).append(hrs)
            except ValueError:
                pass

    totals = {s: 0 for s in _STATUS_ORDER}
    grand_total = 0
    for emp in sorted(agg.keys(), key=lambda x: x.lower()):
        b = agg[emp]
        total = sum(b.get(s, 0) for s in _STATUS_ORDER)
        grand_total += total
        for s in _STATUS_ORDER:
            totals[s] += b.get(s, 0)
        approved = b.get("approved", 0)
        rate = f"{(approved / total * 100):.0f}%" if total else "—"
        samples = cycle.get(emp, [])
        avg_turn = f"{(sum(samples) / len(samples)):.1f}" if samples else "—"
        ps.append([emp, total, b.get("pending", 0), b.get("in_progress", 0),
                   b.get("submitted", 0), approved, rate, avg_turn])

    # Totals row.
    overall_rate = (f"{(totals['approved'] / grand_total * 100):.0f}%"
                    if grand_total else "—")
    ps.append(["All team", grand_total, totals["pending"],
               totals["in_progress"], totals["submitted"], totals["approved"],
               overall_rate, ""])
    totals_row = ps.max_row

    for c in range(1, len(prod_headers) + 1):
        cell = ps.cell(row=header_row, column=c)
        cell.font = header_font
        cell.fill = header_fill
    for c in range(1, len(prod_headers) + 1):
        ps.cell(row=totals_row, column=c).font = Font(bold=True)
    prod_widths = [16, 8, 10, 13, 12, 11, 16, 20]
    for i, w in enumerate(prod_widths, start=1):
        ps.column_dimensions[get_column_letter(i)].width = w

    return wb


@app.get("/api/supervisor/export")
def export_tasks(start: Optional[str] = None,
                 end: Optional[str] = None,
                 user: dict = Depends(require_supervisor)):
    """Export the supervisor's tasks (within an optional YYYY-MM-DD date
    range) as a two-sheet .xlsx workbook: a `Tasks` detail sheet and a
    `Productivity` summary sheet.
    """
    start_iso = _parse_date_bound(start, end=False)
    end_iso = _parse_date_bound(end, end=True)
    if start_iso and end_iso and start_iso > end_iso:
        raise HTTPException(400, "Start date must be on or before end date.")

    tasks = list_tasks_for_supervisor_range(
        user["uid"], start_iso=start_iso, end_iso=end_iso)
    team = _team_employees(user)

    wb = _build_export_workbook(tasks, team, start=start_iso, end=end_iso)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    label = f"{start or 'all'}_to_{end or 'today'}".replace(":", "")
    filename = f"productivity_{label}.xlsx"
    return Response(
        content=buf.getvalue(),
        media_type=("application/vnd.openxmlformats-officedocument"
                    ".spreadsheetml.sheet"),
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
